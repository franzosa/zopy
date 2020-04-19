#!/usr/bin/env python

import os
import sys
import re
import argparse
import csv

from zopy.utils import path2name, iter_rows, die

try:
    import openpyxl as xl
except:
    die( "This script requires the OPENPYXL module" )
    
# argument parsing (python argparse)
parser = argparse.ArgumentParser( ) 
parser.add_argument( "tsv_files", nargs="+", help="" )
parser.add_argument( "--outfile", default=None )
args = parser.parse_args( )

wb = xl.Workbook( )
sheets = []
for i, p in enumerate( args.tsv_files ):
    if i == 0:
        sheets.append( wb.active )
    else:
        sheets.append( wb.create_sheet( ) )
    sheets[-1].title = path2name( p )

for p, ws in zip( args.tsv_files, sheets ):
    for i, row in enumerate( iter_rows( p ) ):
        for j, val in enumerate( row ):
            try:
                val = float( val )
            except:
                pass
            ws.cell( row=i+1, column=j+1, value=val )

if args.outfile is not None:
    outfile = args.outfile
elif len( args.tsv_files ) == 1:
    outfile = "{}.xlsx".format( path2name( args.tsv_files[0] ) )
else:
    outfile = "multisheet.xlsx"           
wb.save( filename=outfile )
